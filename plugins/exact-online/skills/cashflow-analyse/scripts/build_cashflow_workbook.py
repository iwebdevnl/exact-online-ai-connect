#!/usr/bin/env python3
"""Generic cashflow workbook generator (Exact Online, type-based classification).

Usage: python build_cashflow_workbook.py input.json output.xlsx

Input JSON contains the raw movement matrix (mutatie per GLAccount-Type per month);
all accounting and KPIs are computed here and as Excel SUMIFS formulas, so the workbook
is deterministic and fully auditable. No administration-specific account numbers/names.
"""
import json
import sys
from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.utils import get_column_letter

# ---- Platform-wide Exact Online GLAccount.Type map (identical in every administration) ----
TYPE_NL = {
    "10": "Kas", "12": "Bank", "14": "Creditcard", "16": "Payment service (PSP)",
    "20": "Debiteuren", "21": "Vooruitbetaald aan debiteuren", "22": "Crediteuren",
    "24": "Btw", "25": "Personeel te betalen", "26": "Vooruitbetaalde kosten",
    "27": "Nog te betalen kosten", "29": "Vennootschapsbelasting te betalen",
    "40": "Voorraad", "100": "Belasting te betalen", "35": "Cumulatieve afschrijving",
    "90": "Tussen-/verzamelrekening", "30": "Vaste activa", "32": "Overige langlopende activa",
    "50": "Eigen vermogen — kapitaal", "52": "Winstreserve", "55": "Langlopende schuld",
    "60": "Kortlopend deel langlopende schuld", "110": "Omzet", "111": "Kostprijs omzet",
    "120": "Overige kosten", "121": "Verkoop/algemeen/beheer", "122": "Afschrijvingskosten",
    "123": "R&D", "125": "Personeelskosten", "126": "Werkgeverslasten",
    "130": "Bijzondere lasten", "140": "Bijzondere baten", "150": "Belasting over resultaat",
    "160": "Rentebaten/-lasten", "300": "Jaareinde-spiegeling",
    "301": "Jaareinde-kostenverdeling (indirect)", "302": "Jaareinde-kostenverdeling (direct)",
}
LIQUIDE = {"10", "12", "14", "16"}
TECHNISCH = {"300", "301", "302"}
BAL_MAP = {  # balance-account Type -> (categorie, regel)
    "20": ("Operationeel", "Mutatie debiteuren"),
    "21": ("Operationeel", "Mutatie debiteuren"),
    "22": ("Operationeel", "Mutatie crediteuren"),
    "24": ("Operationeel", "Mutatie belastingen (btw e.d.)"),
    "29": ("Operationeel", "Mutatie belastingen (btw e.d.)"),
    "100": ("Operationeel", "Mutatie belastingen (btw e.d.)"),
    "25": ("Operationeel", "Mutatie personeel te betalen"),
    "40": ("Operationeel", "Mutatie voorraad"),
    "26": ("Operationeel", "Mutatie overige werkkapitaal"),
    "27": ("Operationeel", "Mutatie overige werkkapitaal"),
    "90": ("Operationeel", "Mutatie overige werkkapitaal"),
    "35": ("Operationeel", "Afschrijvingen (terugname)"),
    "30": ("Investering", "Investeringen vaste activa"),
    "32": ("Investering", "Overige langlopende activa"),
    "50": ("Financiering", "Mutatie eigen vermogen"),
    "52": ("Financiering", "Mutatie eigen vermogen"),
    "55": ("Financiering", "Mutatie leningen"),
    "60": ("Financiering", "Mutatie leningen"),
}


def classify(t, balance_type):
    """Return (categorie, regel) for a GLAccount Type + BalanceType. Generic, no account numbers."""
    t = str(t)
    if t in LIQUIDE:
        return ("Liquide middelen", "Liquide middelen")
    if t in TECHNISCH:
        return ("Technisch", "Niet-geclassificeerd / technisch")
    if (balance_type or "").upper() == "W":
        return ("Operationeel", "Nettoresultaat")
    return BAL_MAP.get(t, ("Technisch", "Niet-geclassificeerd / technisch"))


# ---- styling ----
FONT = "Arial"
NAVY = "1F3864"
GREY = "F2F2F2"
H1 = Font(name=FONT, size=16, bold=True, color="FFFFFF")
H2 = Font(name=FONT, size=11, bold=True, color="FFFFFF")
BOLD = Font(name=FONT, size=10, bold=True)
NORM = Font(name=FONT, size=10)
ITAL = Font(name=FONT, size=9, italic=True, color="555555")
FILL_NAVY = PatternFill("solid", fgColor=NAVY)
FILL_GREY = PatternFill("solid", fgColor=GREY)
FILL_YELL = PatternFill("solid", fgColor="FFF2CC")
RIGHT = Alignment(horizontal="right")
LEFT = Alignment(horizontal="left", vertical="center")
CTR = Alignment(horizontal="center", vertical="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
EUR = '"€" #,##0;"€" (#,##0);"-"'
EUR2 = '"€" #,##0.00;"€" (#,##0.00);"-"'
DAYS = '0 "dgn";(0) "dgn";"-"'
PCT = '0.0%'
NUMX = '0.0"x"'
MONTHS_NL = {1: "jan", 2: "feb", 3: "mrt", 4: "apr", 5: "mei", 6: "jun",
             7: "jul", 8: "aug", 9: "sep", 10: "okt", 11: "nov", 12: "dec"}


def style_row(ws, row, ncol, font=NORM, fill=None, fmt=None, border=False):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = BORDER
        if fmt and c >= 2:
            cell.number_format = fmt


def main(in_path, out_path):
    with open(in_path) as f:
        d = json.load(f)
    matrix = d.get("matrix", [])
    months = sorted({int(r["maand"]) for r in matrix}) or [1]
    n = len(months)
    begin = float(d.get("beginsaldo_liquide", 0) or 0)
    peil = d.get("peildatum") or date.today().isoformat()
    cur = d.get("valuta", "EUR")

    # ---- Python-side recompute for assertion + KPIs ----
    cat_tot = {"Operationeel": 0.0, "Investering": 0.0, "Financiering": 0.0,
               "Technisch": 0.0, "Liquide middelen": 0.0}
    by_type = {}
    onbekend = set()
    for r in matrix:
        t = str(r["type"])
        cat, _ = classify(t, r.get("balance_type"))
        amt = float(r["mutatie"])
        kas = -amt
        by_type[t] = by_type.get(t, 0.0) + amt
        if cat == "Liquide middelen":
            cat_tot[cat] += amt
        else:
            cat_tot[cat] += kas
        if t not in TYPE_NL:
            onbekend.add(t)
    netto = cat_tot["Operationeel"] + cat_tot["Investering"] + cat_tot["Financiering"] + cat_tot["Technisch"]
    mut_liq = cat_tot["Liquide middelen"]
    recon_diff = round(netto - mut_liq, 2)
    eind = begin + netto

    wb = Workbook()

    # ================= BRONDATA =================
    bd = wb.active
    bd.title = "Brondata"
    hdr = ["Type", "Omschrijving", "BalanceType", "Maand", "Mutatie (AmountDC)", "Regel", "Categorie", "Kaseffect"]
    bd.append(hdr)
    style_row(bd, 1, len(hdr), font=H2, fill=FILL_NAVY)
    rownum = 1
    for r in matrix:
        rownum += 1
        t = str(r["type"])
        cat, regel = classify(t, r.get("balance_type"))
        oms = r.get("omschrijving") or TYPE_NL.get(t, "Onbekend type")
        bd.append([t, oms, (r.get("balance_type") or ""), int(r["maand"]), float(r["mutatie"]), regel, cat, None])
        bd.cell(row=rownum, column=8).value = f"=-E{rownum}"
        bd.cell(row=rownum, column=5).number_format = EUR2
        bd.cell(row=rownum, column=8).number_format = EUR2
    last = rownum
    for col, w in zip("ABCDEFGH", (8, 26, 12, 8, 18, 26, 16, 16)):
        bd.column_dimensions[col].width = w
    bd.freeze_panes = "A2"
    # named full-column ranges used by SUMIFS
    TY, MA, KA, RE, CA, MU = "Brondata!$A:$A", "Brondata!$D:$D", "Brondata!$H:$H", "Brondata!$F:$F", "Brondata!$G:$G", "Brondata!$E:$E"

    def sif_regel(regel, m):
        return f'=SUMIFS({KA},{RE},"{regel}",{MA},{m})'

    def sif_regel_tot(regel):
        return f'=SUMIFS({KA},{RE},"{regel}")'

    def sif_cat(cat, m):
        return f'=SUMIFS({KA},{CA},"{cat}",{MA},{m})'

    def sif_cat_tot(cat):
        return f'=SUMIFS({KA},{CA},"{cat}")'

    def sif_type(t, m):
        return f'=SUMIFS({KA},{TY},"{t}",{MA},{m})'

    # ================= CLASSIFICATIE =================
    cl = wb.create_sheet("Classificatie")
    cl.append(["Type", "Omschrijving (NL)", "Categorie", "Regel in overzicht"])
    style_row(cl, 1, 4, font=H2, fill=FILL_NAVY)
    for t in sorted(TYPE_NL, key=lambda x: int(x)):
        bt = "W" if t in {"110", "111", "120", "121", "122", "123", "125", "126", "130", "140", "150", "160"} else ("B/W" if t in TECHNISCH else "B")
        cat, regel = classify(t, "W" if bt == "W" else "B")
        cl.append([t, TYPE_NL[t], cat, regel])
    for col, w in zip("ABCD", (8, 34, 18, 30)):
        cl.column_dimensions[col].width = w
    cl.freeze_panes = "A2"
    note = cl.cell(row=cl.max_row + 2, column=1,
                   value="Classificatie op platform-breed rekeningtype (Exact Online), niet op rekeningnummer of -naam. "
                         "Alle W&V-rekeningen vormen samen het nettoresultaat; balansrekeningen worden op Type ingedeeld.")
    note.font = ITAL
    cl.merge_cells(start_row=note.row, start_column=1, end_row=note.row, end_column=4)
    note.alignment = WRAP

    # helper to lay out a month-columned statement
    def month_header(ws, row, first_label):
        ws.cell(row=row, column=1, value=first_label)
        for i, m in enumerate(months):
            ws.cell(row=row, column=2 + i, value=MONTHS_NL.get(m, str(m)))
        ws.cell(row=row, column=2 + n, value="Totaal")
        style_row(ws, row, 2 + n, font=H2, fill=FILL_NAVY)
        ws.cell(row=row, column=1).alignment = LEFT

    def line(ws, row, label, formula_fn, total_fn, bold=False, fill=None, fmt=EUR):
        ws.cell(row=row, column=1, value=label).font = BOLD if bold else NORM
        for i, m in enumerate(months):
            c = ws.cell(row=row, column=2 + i, value=formula_fn(m))
            c.number_format = fmt
            c.font = BOLD if bold else NORM
        tc = ws.cell(row=row, column=2 + n, value=total_fn())
        tc.number_format = fmt
        tc.font = BOLD if bold else NORM
        if fill:
            style_row(ws, row, 2 + n, font=BOLD if bold else NORM, fill=fill, fmt=fmt)
            ws.cell(row=row, column=1).alignment = LEFT

    def colsum(ws, row, rows_to_sum):
        """write a subtotal row summing given statement rows, per column."""
        for i in range(n + 1):
            col = get_column_letter(2 + i)
            parts = "+".join(f"{col}{rr}" for rr in rows_to_sum)
            ws.cell(row=row, column=2 + i, value=f"={parts}")

    # ================= KASSTROOM (INDIRECT) =================
    ind = wb.create_sheet("Kasstroom (indirect)")
    ind.cell(row=1, column=1, value=f"Kasstroomoverzicht — indirecte methode  ({d.get('periode_label','')})").font = Font(name=FONT, size=13, bold=True, color=NAVY)
    month_header(ind, 3, f"Bedragen in {cur}")
    op_lines = ["Nettoresultaat", "Afschrijvingen (terugname)", "Mutatie debiteuren",
                "Mutatie voorraad", "Mutatie crediteuren", "Mutatie belastingen (btw e.d.)",
                "Mutatie personeel te betalen", "Mutatie overige werkkapitaal"]
    inv_lines = ["Investeringen vaste activa", "Overige langlopende activa"]
    fin_lines = ["Mutatie eigen vermogen", "Mutatie leningen"]
    r = 4
    op_rows = []
    for lab in op_lines:
        line(ind, r, lab, lambda m, L=lab: sif_regel(L, m), lambda L=lab: sif_regel_tot(L))
        op_rows.append(r)
        r += 1
    ind.cell(row=r, column=1, value="Operationele kasstroom")
    colsum(ind, r, op_rows)
    style_row(ind, r, 2 + n, font=BOLD, fill=FILL_GREY, fmt=EUR)
    ind.cell(row=r, column=1).font = BOLD
    row_op = r
    r += 2
    inv_rows = []
    for lab in inv_lines:
        line(ind, r, lab, lambda m, L=lab: sif_regel(L, m), lambda L=lab: sif_regel_tot(L))
        inv_rows.append(r)
        r += 1
    ind.cell(row=r, column=1, value="Investeringskasstroom")
    colsum(ind, r, inv_rows)
    style_row(ind, r, 2 + n, font=BOLD, fill=FILL_GREY, fmt=EUR)
    ind.cell(row=r, column=1).font = BOLD
    row_inv = r
    r += 2
    fin_rows = []
    for lab in fin_lines:
        line(ind, r, lab, lambda m, L=lab: sif_regel(L, m), lambda L=lab: sif_regel_tot(L))
        fin_rows.append(r)
        r += 1
    ind.cell(row=r, column=1, value="Financieringskasstroom")
    colsum(ind, r, fin_rows)
    style_row(ind, r, 2 + n, font=BOLD, fill=FILL_GREY, fmt=EUR)
    ind.cell(row=r, column=1).font = BOLD
    row_fin = r
    r += 2
    line(ind, r, "Niet-geclassificeerd / technisch",
         lambda m: sif_regel("Niet-geclassificeerd / technisch", m),
         lambda: sif_regel_tot("Niet-geclassificeerd / technisch"))
    row_tech = r
    r += 1
    ind.cell(row=r, column=1, value="Netto kasstroom").font = BOLD
    colsum(ind, r, [row_op, row_inv, row_fin, row_tech])
    style_row(ind, r, 2 + n, font=BOLD, fill=FILL_NAVY, fmt=EUR)
    for c in range(1, 2 + n + 1):
        ind.cell(row=r, column=c).font = H2
    row_netto = r
    r += 2
    # begin/eind balances (running)
    ind.cell(row=r, column=1, value="Beginsaldo liquide middelen").font = NORM
    ind.cell(row=r, column=2, value=begin).number_format = EUR
    for i in range(1, n):
        ind.cell(row=r, column=2 + i, value=f"={get_column_letter(2 + i - 1)}{r + 1}").number_format = EUR
    ind.cell(row=r, column=2 + n, value=begin).number_format = EUR
    row_begin = r
    r += 1
    ind.cell(row=r, column=1, value="Eindsaldo liquide middelen").font = BOLD
    for i in range(n):
        col = get_column_letter(2 + i)
        ind.cell(row=r, column=2 + i, value=f"={col}{row_begin}+{col}{row_netto}").number_format = EUR
    ind.cell(row=r, column=2 + n, value=f"={get_column_letter(1 + n)}{r}").number_format = EUR
    style_row(ind, r, 2 + n, font=BOLD, fill=FILL_GREY, fmt=EUR)
    ind.cell(row=r, column=1).font = BOLD
    row_eind = r
    r += 2
    # control: actual liquide movement vs netto
    ind.cell(row=r, column=1, value="Controle — werkelijke mutatie liquide middelen").font = ITAL
    for i, m in enumerate(months):
        ind.cell(row=r, column=2 + i, value=f'=SUMIFS({MU},{CA},"Liquide middelen",{MA},{m})').number_format = EUR
    ind.cell(row=r, column=2 + n, value=f'=SUMIFS({MU},{CA},"Liquide middelen")').number_format = EUR
    row_ctrl = r
    r += 1
    ind.cell(row=r, column=1, value="Controle — verschil (moet 0 zijn)").font = ITAL
    for i in range(n + 1):
        col = get_column_letter(2 + i)
        ind.cell(row=r, column=2 + i, value=f"={col}{row_netto}-{col}{row_ctrl}").number_format = EUR2
    style_row(ind, r, 2 + n, font=ITAL, fill=FILL_YELL, fmt=EUR2)
    ind.cell(row=r, column=1).font = ITAL
    ind.column_dimensions["A"].width = 38
    for i in range(n + 1):
        ind.column_dimensions[get_column_letter(2 + i)].width = 13
    ind.freeze_panes = "B4"

    # ================= KASSTROOM (DIRECT, AFGELEID) =================
    dr = wb.create_sheet("Kasstroom (direct)")
    dr.cell(row=1, column=1, value=f"Kasstroomoverzicht — directe methode (afgeleid)  ({d.get('periode_label','')})").font = Font(name=FONT, size=13, bold=True, color=NAVY)
    month_header(dr, 3, f"Bedragen in {cur}")

    def f_klanten(m):
        return f'=SUMIFS({KA},{TY},"110",{MA},{m})+SUMIFS({KA},{RE},"Mutatie debiteuren",{MA},{m})'

    def f_lev(m):
        return (f'=SUMIFS({KA},{TY},"111",{MA},{m})+SUMIFS({KA},{RE},"Mutatie crediteuren",{MA},{m})'
                f'+SUMIFS({KA},{RE},"Mutatie voorraad",{MA},{m})')

    def f_pers(m):
        return f'=SUMIFS({KA},{TY},"125",{MA},{m})+SUMIFS({KA},{TY},"126",{MA},{m})+SUMIFS({KA},{RE},"Mutatie personeel te betalen",{MA},{m})'

    def f_overig(m):
        return (f'=SUMIFS({KA},{TY},"120",{MA},{m})+SUMIFS({KA},{TY},"121",{MA},{m})'
                f'+SUMIFS({KA},{TY},"123",{MA},{m})+SUMIFS({KA},{TY},"130",{MA},{m})'
                f'+SUMIFS({KA},{TY},"140",{MA},{m})+SUMIFS({KA},{RE},"Mutatie overige werkkapitaal",{MA},{m})')

    def f_belas(m):
        return f'=SUMIFS({KA},{TY},"150",{MA},{m})+SUMIFS({KA},{RE},"Mutatie belastingen (btw e.d.)",{MA},{m})'

    def f_rente(m):
        return f'=SUMIFS({KA},{TY},"160",{MA},{m})'

    direct_specs = [
        ("Ontvangen van klanten", f_klanten),
        ("Betaald aan leveranciers en voorraad", f_lev),
        ("Betaald aan personeel", f_pers),
        ("Overige operationele uitgaven", f_overig),
        ("Betaalde belastingen", f_belas),
        ("Ontvangen/betaalde rente", f_rente),
    ]
    rr = 4
    drows = []
    for lab, fn in direct_specs:
        for i, m in enumerate(months):
            dr.cell(row=rr, column=2 + i, value=fn(m)).number_format = EUR
        # total via no-month variant: rebuild by summing months
        cols = "+".join(f"{get_column_letter(2 + i)}{rr}" for i in range(n))
        dr.cell(row=rr, column=2 + n, value=f"={cols}").number_format = EUR
        dr.cell(row=rr, column=1, value=lab).font = NORM
        drows.append(rr)
        rr += 1
    # plug line so direct OCF == indirect OCF exactly
    dr.cell(row=rr, column=1, value="Overige operationele posten / non-cash (saldo)").font = ITAL
    for i, m in enumerate(months):
        col = get_column_letter(2 + i)
        named = "+".join(f"{col}{x}" for x in drows)
        dr.cell(row=rr, column=2 + i, value=f'=SUMIFS({KA},{CA},"Operationeel",{MA},{m})-({named})').number_format = EUR
    coln = get_column_letter(2 + n)
    named_t = "+".join(f"{coln}{x}" for x in drows)
    dr.cell(row=rr, column=2 + n, value=f'=SUMIFS({KA},{CA},"Operationeel")-({named_t})').number_format = EUR
    plug = rr
    rr += 1
    dr.cell(row=rr, column=1, value="Operationele kasstroom (direct)").font = BOLD
    colsum(dr, rr, drows + [plug])
    style_row(dr, rr, 2 + n, font=BOLD, fill=FILL_GREY, fmt=EUR)
    dr.cell(row=rr, column=1).font = BOLD
    row_dir_ocf = rr
    rr += 1
    dr.cell(row=rr, column=1, value="Controle — verschil met indirecte methode (moet 0 zijn)").font = ITAL
    for i in range(n + 1):
        col = get_column_letter(2 + i)
        dr.cell(row=rr, column=2 + i, value=f"={col}{row_dir_ocf}-'Kasstroom (indirect)'!{col}{row_op}").number_format = EUR2
    style_row(dr, rr, 2 + n, font=ITAL, fill=FILL_YELL, fmt=EUR2)
    dr.cell(row=rr, column=1).font = ITAL
    note2 = dr.cell(row=rr + 2, column=1,
                    value="Afgeleide directe methode: ontvangsten en uitgaven volgen uit de resultaatposten, gecorrigeerd voor "
                          "de bijbehorende werkkapitaalmutaties. Sluit per definitie aan op de operationele kasstroom van de indirecte methode.")
    note2.font = ITAL
    dr.merge_cells(start_row=note2.row, start_column=1, end_row=note2.row, end_column=2 + n)
    note2.alignment = WRAP
    dr.column_dimensions["A"].width = 40
    for i in range(n + 1):
        dr.column_dimensions[get_column_letter(2 + i)].width = 13
    dr.freeze_panes = "B4"

    # ================= GRAFIEKDATA =================
    gd = wb.create_sheet("Grafiekdata")
    gd.append(["Maand", "Operationeel", "Investering", "Financiering", "Netto", "Eindsaldo"])
    style_row(gd, 1, 6, font=H2, fill=FILL_NAVY)
    for i, m in enumerate(months):
        gr = i + 2
        gd.cell(row=gr, column=1, value=MONTHS_NL.get(m, str(m)))
        gd.cell(row=gr, column=2, value=sif_cat("Operationeel", m)).number_format = EUR
        gd.cell(row=gr, column=3, value=sif_cat("Investering", m)).number_format = EUR
        gd.cell(row=gr, column=4, value=sif_cat("Financiering", m)).number_format = EUR
        gd.cell(row=gr, column=5, value=f"=B{gr}+C{gr}+D{gr}+SUMIFS({KA},{CA},\"Technisch\",{MA},{m})").number_format = EUR
        if i == 0:
            gd.cell(row=gr, column=6, value=f"={begin}+E{gr}").number_format = EUR
        else:
            gd.cell(row=gr, column=6, value=f"=F{gr-1}+E{gr}").number_format = EUR
    gd_last = n + 1
    for col, w in zip("ABCDEF", (8, 14, 14, 14, 14, 14)):
        gd.column_dimensions[col].width = w

    # ================= WERKKAPITAAL & KPI'S =================
    kp = wb.create_sheet("Werkkapitaal & KPI's")
    deb = d.get("debiteuren", []) or []
    cred = d.get("crediteuren", []) or []
    open_deb = float(d["open_debiteuren_totaal"]) if d.get("open_debiteuren_totaal") is not None else sum(float(x["bedrag"]) for x in deb)
    open_cred = float(d["open_crediteuren_totaal"]) if d.get("open_crediteuren_totaal") is not None else sum(float(x["bedrag"]) for x in cred)
    voorraad_eind = d.get("voorraad_eindstand")
    kp.cell(row=1, column=1, value="Werkkapitaal & kasstroom-KPI's").font = Font(name=FONT, size=13, bold=True, color=NAVY)
    kp.cell(row=2, column=1, value=f"Periode: {d.get('periode_label','')} — {n} maand(en). Cijfers indicatief, stand per " + peil).font = ITAL
    # inputs block
    inp = [
        ("Maanden in periode", n, '0'),
        ("Omzet (periode)", f'=SUMIFS({KA},{TY},"110")', EUR),
        ("Kostprijs + overige kosten (periode)", f'=-(SUMIFS({KA},{TY},"111")+SUMIFS({KA},{TY},"120")+SUMIFS({KA},{TY},"121")+SUMIFS({KA},{TY},"123"))', EUR),
        ("Open debiteuren (stand)", open_deb, EUR),
        ("Open crediteuren (stand)", open_cred, EUR),
        ("Voorraad (eindstand)", (float(voorraad_eind) if voorraad_eind is not None else "n.v.t."), EUR),
        ("Operationele kasstroom (OCF)", f"='Kasstroom (indirect)'!{get_column_letter(2+n)}{row_op}", EUR),
        ("Investeringskasstroom", f"='Kasstroom (indirect)'!{get_column_letter(2+n)}{row_inv}", EUR),
        ("Liquide eindsaldo", f"='Kasstroom (indirect)'!{get_column_letter(2+n)}{row_eind}", EUR),
    ]
    rr = 4
    kp.cell(row=rr, column=1, value="Grondslagen").font = H2
    kp.cell(row=rr, column=1).fill = FILL_NAVY
    kp.cell(row=rr, column=2).fill = FILL_NAVY
    rr += 1
    ref = {}
    for lab, val, fmt in inp:
        kp.cell(row=rr, column=1, value=lab).font = NORM
        c = kp.cell(row=rr, column=2, value=val)
        c.number_format = fmt
        c.font = NORM
        ref[lab] = f"B{rr}"
        rr += 1
    rr += 1
    kp.cell(row=rr, column=1, value="Kengetal").font = H2
    kp.cell(row=rr, column=2, value="Waarde").font = H2
    kp.cell(row=rr, column=3, value="Toelichting").font = H2
    style_row(kp, rr, 3, font=H2, fill=FILL_NAVY)
    rr += 1
    M = ref["Maanden in periode"]
    omz = ref["Omzet (periode)"]
    kost = ref["Kostprijs + overige kosten (periode)"]
    od = ref["Open debiteuren (stand)"]
    oc = ref["Open crediteuren (stand)"]
    vr = ref["Voorraad (eindstand)"]
    ocf = ref["Operationele kasstroom (OCF)"]
    inv = ref["Investeringskasstroom"]
    liq = ref["Liquide eindsaldo"]
    kpis = [
        ("DSO — debiteurendagen", f'=IFERROR({od}/({omz}/{M}*12)*365,"-")', DAYS, "Gem. dagen tot inning"),
        ("DPO — crediteurendagen", f'=IFERROR({oc}/({kost}/{M}*12)*365,"-")', DAYS, "Gem. dagen tot betaling"),
        ("DIO — voorraaddagen", (f'=IFERROR({vr}/({kost}/{M}*12)*365,"-")' if voorraad_eind is not None else '="-"'), DAYS, "Gem. dagen voorraad"),
        ("Kasconversiecyclus (CCC)", (f'=IFERROR(B{rr}+B{rr+2}-B{rr+1},"-")' if voorraad_eind is not None else f'=IFERROR(B{rr}-B{rr+1},"-")'), DAYS, "DSO + DIO − DPO"),
        ("Vrije kasstroom (FCF)", f"={ocf}+{inv}", EUR, "OCF − investeringen"),
        ("Kasstroommarge", f'=IFERROR({ocf}/{omz},"-")', PCT, "OCF / omzet"),
        ("Operationele-kasstroomratio", f'=IFERROR({ocf}/{oc},"-")', NUMX, "OCF / korte schulden (proxy: crediteuren)"),
        ("Cash runway (mnd)", f'=IFERROR(IF({ocf}<0,{liq}/(-{ocf}/{M}),"ruim"),"-")', '0.0', "Maanden bij huidige netto-uitstroom"),
    ]
    kpi_start = rr
    for lab, val, fmt, toel in kpis:
        kp.cell(row=rr, column=1, value=lab).font = BOLD
        c = kp.cell(row=rr, column=2, value=val)
        c.number_format = fmt
        c.font = BOLD
        kp.cell(row=rr, column=3, value=toel).font = ITAL
        rr += 1
    kp.column_dimensions["A"].width = 34
    kp.column_dimensions["B"].width = 16
    kp.column_dimensions["C"].width = 42

    # ================= DASHBOARD =================
    db = wb.create_sheet("Dashboard")
    wb.move_sheet("Dashboard", -(len(wb.sheetnames) - 1))
    db.sheet_view.showGridLines = False
    db.merge_cells("A1:H1")
    t = db.cell(row=1, column=1, value=f"Cashflow-analyse — {d.get('bedrijf','')}")
    t.font = H1
    t.fill = FILL_NAVY
    t.alignment = LEFT
    db.row_dimensions[1].height = 28
    db.cell(row=2, column=1, value=f"{d.get('periode_label','')}   ·   methode: direct + indirect   ·   stand per {peil}").font = ITAL
    cards = [
        ("Liquide eindsaldo", f"='Kasstroom (indirect)'!{get_column_letter(2+n)}{row_eind}", EUR),
        ("Operationele kasstroom", f"='Kasstroom (indirect)'!{get_column_letter(2+n)}{row_op}", EUR),
        ("Vrije kasstroom", f"='Werkkapitaal & KPI''s'!B{kpi_start+4}", EUR),
        ("Netto kasstroom periode", f"='Kasstroom (indirect)'!{get_column_letter(2+n)}{row_netto}", EUR),
        ("DSO", f"='Werkkapitaal & KPI''s'!B{kpi_start}", DAYS),
        ("Kasconversiecyclus", f"='Werkkapitaal & KPI''s'!B{kpi_start+3}", DAYS),
    ]
    cr0 = 4
    for i, (lab, val, fmt) in enumerate(cards):
        col = 1 + (i % 3) * 3
        row = cr0 + (i // 3) * 3
        lc = db.cell(row=row, column=col, value=lab)
        lc.font = Font(name=FONT, size=9, bold=True, color="FFFFFF")
        lc.fill = FILL_NAVY
        lc.alignment = LEFT
        db.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
        vc = db.cell(row=row + 1, column=col, value=val)
        vc.number_format = fmt
        vc.font = Font(name=FONT, size=15, bold=True, color=NAVY)
        vc.fill = FILL_GREY
        vc.alignment = LEFT
        db.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 1)
    for c in "ABCDEFGH":
        db.column_dimensions[c].width = 13

    # charts
    cats_ref = Reference(gd, min_col=1, min_row=2, max_row=gd_last)
    line_ch = LineChart()
    line_ch.title = "Verloop liquide middelen"
    line_ch.height = 7.5
    line_ch.width = 16
    data = Reference(gd, min_col=6, min_row=1, max_row=gd_last)
    line_ch.add_data(data, titles_from_data=True)
    line_ch.set_categories(cats_ref)
    line_ch.y_axis.numFmt = '#,##0'
    line_ch.y_axis.majorGridlines = None
    db.add_chart(line_ch, "A11")

    bar_net = BarChart()
    bar_net.type = "col"
    bar_net.title = "Netto kasstroom per maand"
    bar_net.height = 7.5
    bar_net.width = 16
    bn = Reference(gd, min_col=5, min_row=1, max_row=gd_last)
    bar_net.add_data(bn, titles_from_data=True)
    bar_net.set_categories(cats_ref)
    bar_net.y_axis.numFmt = '#,##0'
    db.add_chart(bar_net, "E11")

    bar_act = BarChart()
    bar_act.type = "col"
    bar_act.grouping = "stacked"
    bar_act.overlap = 100
    bar_act.title = "Kasstroom per activiteit per maand"
    bar_act.height = 7.5
    bar_act.width = 16
    ba = Reference(gd, min_col=2, max_col=4, min_row=1, max_row=gd_last)
    bar_act.add_data(ba, titles_from_data=True)
    bar_act.set_categories(cats_ref)
    bar_act.y_axis.numFmt = '#,##0'
    db.add_chart(bar_act, "A27")

    # ================= TOELICHTING =================
    to = wb.create_sheet("Toelichting")
    wb.move_sheet("Toelichting", -(len(wb.sheetnames) - 1))
    to.sheet_view.showGridLines = False
    to.column_dimensions["A"].width = 100
    blocks = [
        (f"Cashflow-analyse — {d.get('bedrijf','')}", H1, FILL_NAVY),
        (f"Administratie {d.get('administratie','')} · periode {d.get('periode_label','')} · valuta {cur}", ITAL, None),
        ("", NORM, None),
        ("Opzet", BOLD, None),
        ("Dit overzicht toont de kasstroom volgens zowel de directe (afgeleide) als de indirecte methode, "
         "verdeeld in operationele, investerings- en financieringskasstroom. De analyse is generiek: rekeningen "
         "worden geclassificeerd op hun platform-brede rekeningtype in Exact Online, niet op administratie-specifieke "
         "grootboeknummers of -namen.", NORM, None),
        ("Rekenwijze", BOLD, None),
        ("Per boeking geldt debet = credit, dus over alle rekeningen samen is de mutatie nul. De mutatie op de "
         "liquide rekeningen is daarmee gelijk aan minus de mutatie op alle overige rekeningen — de basis van de "
         "indirecte methode. Het kaseffect van elke niet-liquide rekening is daarom: − mutatie. Hierdoor sluiten beide "
         "methoden per definitie op elkaar aan (zie de controleregels onderaan elk tabblad).", NORM, None),
        ("Tabbladen", BOLD, None),
        ("· Dashboard — kerncijfers en grafieken\n"
         "· Kasstroom (indirect) — resultaat + afschrijvingen + werkkapitaal + investering + financiering\n"
         "· Kasstroom (direct) — ontvangsten en uitgaven per categorie (afgeleid)\n"
         "· Werkkapitaal & KPI's — DSO, DPO, DIO, kasconversiecyclus, vrije kasstroom, marges\n"
         "· Brondata — de ruwe mutatie per rekeningtype per maand (bron van alle formules)\n"
         "· Classificatie — de gebruikte rekeningtype-indeling", NORM, None),
        ("Aandachtspunten", BOLD, None),
        (("· Het beginsaldo liquide middelen is " + ("bevestigd." if d.get("beginsaldo_bevestigd") else "NIET bevestigd — controleer dit tegen de werkelijke bankstand.") + "\n"
          "· Cijfers zijn indicatief en betreffen de stand per " + peil + "; ook afgesloten periodes kunnen nog corrigeren.\n"
          "· Onbekende rekeningtypes" + (": " + ", ".join(sorted(onbekend)) + " — controleer de regel 'Niet-geclassificeerd / technisch'." if onbekend else " zijn niet aangetroffen.")), NORM, FILL_YELL),
    ]
    rr = 1
    for text, font, fill in blocks:
        c = to.cell(row=rr, column=1, value=text)
        c.font = font
        c.alignment = WRAP
        if fill:
            c.fill = fill
        if text and "\n" in text:
            to.row_dimensions[rr].height = 15 * (text.count("\n") + 1)
        elif font == H1:
            to.row_dimensions[rr].height = 26
        rr += 1

    # ---- optional debiteuren / crediteuren ----
    def aging_sheet(name, items, kind):
        if not items:
            return
        sh = wb.create_sheet(name)
        sh.append([kind, "Bedrag", "Vervaldatum", "Dagen vervallen", "Ouderdom"])
        style_row(sh, 1, 5, font=H2, fill=FILL_NAVY)
        peil_d = datetime.fromisoformat(peil).date()
        for it in items:
            try:
                vd = datetime.fromisoformat(str(it.get("vervaldatum"))[:10]).date()
                dgn = (peil_d - vd).days
            except Exception:
                dgn = ""
            if dgn == "":
                bucket = "onbekend"
            elif dgn <= 0:
                bucket = "niet vervallen"
            elif dgn <= 30:
                bucket = "1-30"
            elif dgn <= 60:
                bucket = "31-60"
            elif dgn <= 90:
                bucket = "61-90"
            else:
                bucket = ">90"
            sh.append([it.get("naam", ""), float(it["bedrag"]), str(it.get("vervaldatum", "")), dgn, bucket])
            sh.cell(row=sh.max_row, column=2).number_format = EUR
        last_data = sh.max_row
        tot = last_data + 1
        sh.cell(row=tot, column=1, value="Totaal").font = BOLD
        sh.cell(row=tot, column=2, value=f"=SUM(B2:B{last_data})").number_format = EUR
        sh.cell(row=tot, column=2).font = BOLD
        for col, w in zip("ABCDE", (34, 16, 14, 16, 14)):
            sh.column_dimensions[col].width = w
        sh.freeze_panes = "A2"

    aging_sheet("Debiteuren", deb, "Debiteur")
    aging_sheet("Crediteuren", cred, "Crediteur")

    wb.save(out_path)
    summary = {
        "recon_ok": abs(recon_diff) < 1.0,
        "verschil_netto_vs_liquide": recon_diff,
        "operationeel": round(cat_tot["Operationeel"], 2),
        "investering": round(cat_tot["Investering"], 2),
        "financiering": round(cat_tot["Financiering"], 2),
        "technisch_rest": round(cat_tot["Technisch"], 2),
        "netto_kasstroom": round(netto, 2),
        "mutatie_liquide": round(mut_liq, 2),
        "beginsaldo": round(begin, 2),
        "eindsaldo": round(eind, 2),
        "maanden": months,
        "onbekende_types": sorted(onbekend),
        "output": out_path,
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    if len(sys.argv) != 3:
        print("Usage: python build_cashflow_workbook.py input.json output.xlsx")
        sys.exit(1)
    main(sys.argv[1], sys.argv[2])
